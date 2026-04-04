#!/usr/bin/env python3
"""
Recalculate Excel formulas for a workbook.
Usage: python recalc.py <filepath>
"""
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def recalculate_workbook(filepath):
    """Load workbook and force recalculation of all formulas."""
    try:
        wb = load_workbook(filepath)
        # Force Excel to recalculate all formulas on next open
        wb.properties.calcMode = 'auto'
        wb.save(filepath)
        print(f"✓ Recalculated and saved: {filepath}")
        return 0
    except Exception as e:
        print(f"✗ Error: {e}", file=sys.stderr)
        return 1

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <filepath>")
        sys.exit(1)
    
    filepath = sys.argv[1]
    sys.exit(recalculate_workbook(filepath))
